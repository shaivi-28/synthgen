"""
NFS Daily Settlement Statement Generator
Generates an Excel settlement file matching the format NFS sends to the bank.

Format derived from: NTSLIDF010326_1C.xlsx
Structure:
  - Header: "Daily Settlement Statement for <BANK> as on DD/MM/YYYY"
  - Column headers: Description | No of Txns | Debit | Credit
  - Line items grouped by: Acquirer (outgoing) and Issuer (incoming) sections
  - Subtotals, Settlement Amount, Net Adjusted, Final Settlement Amount
  - Dispute Adjustments section at the end

Fee Rates (verified from real NFS settlement file):
  Issuer WDL Approved Fee:           ₹19.00 per txn
  Issuer WDL Approved Fee GST:       ₹3.42  per txn (18%)
  Issuer NPCI Switching Fee (WDL):   ₹0.30  per txn
  Issuer NPCI Switching GST:         ₹0.054 per txn (18%)
  Issuer BI Approved Fee:            ₹7.00  per txn
  Issuer BI Approved Fee GST:        ₹1.26  per txn (18%)
  Issuer MS Approved Fee:            ₹7.00  per txn
  Issuer PC Approved Fee:            ₹7.00  per txn
  Late Reversal fee refund:          ₹19.00 per reversed txn (credit back)
"""

from pathlib import Path
from datetime import datetime
from typing import Optional
from dataclasses import dataclass, field
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# FEE RATES  (from real NFS settlement file)
# ─────────────────────────────────────────────────────────────
FEE = {
    "wdl_fee":          19.00,
    "wdl_fee_gst":       3.42,   # 18% of 19
    "npci_switch":       0.30,
    "npci_switch_gst":   0.054,  # 18% of 0.30
    "bi_fee":            7.00,
    "bi_fee_gst":        1.26,   # 18% of 7
    "ms_fee":            7.00,
    "ms_fee_gst":        1.26,
    "pc_fee":            7.00,
    "pc_fee_gst":        1.26,
    "reversal_fee":     19.00,   # refunded on late reversal
    "reversal_fee_gst":  3.42,
}

GST_RATE = 0.18


# ─────────────────────────────────────────────────────────────
# STATS AGGREGATOR  (populated from generated transactions)
# ─────────────────────────────────────────────────────────────

@dataclass
class SettlementStats:
    """Aggregated counts and amounts built from generated transaction groups."""
    bank_name: str = "TEST BANK"
    settlement_date: Optional[datetime] = None

    # Issuer counts (bank's own cardholders using other banks' ATMs)
    issuer_wdl_approved:     int   = 0
    issuer_wdl_declined:     int   = 0
    issuer_wdl_total_amount: float = 0.0   # sum of all withdrawal amounts in INR

    issuer_bi_approved:      int   = 0
    issuer_bi_declined:      int   = 0

    issuer_ms_approved:      int   = 0
    issuer_ms_declined:      int   = 0

    issuer_pc_approved:      int   = 0
    issuer_pc_declined:      int   = 0

    # Late reversals (transactions reversed after settlement)
    issuer_late_reversals:   int   = 0
    issuer_late_rev_amount:  float = 0.0

    # Dispute adjustments
    chargeback_count:        int   = 0
    chargeback_debit:        float = 0.0
    chargeback_credit:       float = 0.0


def stats_from_manifest(manifest: dict, bank_name: str) -> SettlementStats:
    """
    Build SettlementStats from the matrix generator manifest.
    Only counts rows that represent actual transactions (not disputes/chargebacks).

    Logic:
      - case value NFS=1: bank is ISSUER (bank owes NFS for txn)
      - transaction type W1/WDL: withdrawal
      - transaction type BI: balance inquiry
      - transaction type MS: mini statement
      - transaction type PC: PIN change
      - resp_code != 00: declined
      - msg_type 0420: reversal
    """
    stats = SettlementStats(bank_name=bank_name)

    # Parse settlement date from manifest
    date_str = manifest.get("tran_date", "")
    if date_str:
        try:
            stats.settlement_date = datetime.strptime(date_str, "%d%m%Y")
        except ValueError:
            stats.settlement_date = datetime.today()

    for row in manifest.get("rows", []):
        nfs_val  = row.get("nfs_value")
        is_ok    = row.get("is_ok", False)
        scenario = row.get("scenario_id", "")
        amount   = row.get("amount_inr", 0.0)
        # variant  = row.get("variant", "none")  # for future use

        # Only count NFS=1 rows (issuer side) for settlement
        if nfs_val != 1:
            continue

        # Derive transaction type from scenario_id
        ttype = "WDL"   # default
        if "balance_enquiry" in scenario or "bi" in scenario:
            ttype = "BI"
        elif "mini_statement" in scenario:
            ttype = "MS"
        elif "pin_change" in scenario:
            ttype = "PC"

        # Reversed or declined
        is_reversal = "reversal" in scenario or "reversed" in scenario
        is_declined = row.get("nfs_value") == -1 or "declined" in scenario or "decline" in scenario

        if is_reversal:
            stats.issuer_late_reversals += 1
            stats.issuer_late_rev_amount += amount
        elif is_declined or row.get("resp_code", "00") not in ("00", "000", "0"):
            if ttype == "WDL":
                stats.issuer_wdl_declined += 1
            elif ttype == "BI":
                stats.issuer_bi_declined += 1
            elif ttype == "MS":
                stats.issuer_ms_declined += 1
            elif ttype == "PC":
                stats.issuer_pc_declined += 1
        else:
            if ttype == "WDL":
                stats.issuer_wdl_approved += 1
                stats.issuer_wdl_total_amount += amount
            elif ttype == "BI":
                stats.issuer_bi_approved += 1
            elif ttype == "MS":
                stats.issuer_ms_approved += 1
            elif ttype == "PC":
                stats.issuer_pc_approved += 1

    return stats


# ─────────────────────────────────────────────────────────────
# CALCULATED LINE ITEMS
# ─────────────────────────────────────────────────────────────

def compute_line_items(s: SettlementStats) -> dict:
    """
    Compute all monetary values for every line in the settlement statement.
    Returns a dict keyed by description → (count, debit, credit).
    Debit = bank pays NFS. Credit = NFS pays bank.
    """
    f = FEE
    n = s   # shorthand

    wdl_fee       = round(n.issuer_wdl_approved * f["wdl_fee"], 2)
    wdl_fee_gst   = round(n.issuer_wdl_approved * f["wdl_fee_gst"], 2)
    wdl_switch    = round(n.issuer_wdl_approved * f["npci_switch"], 2)
    wdl_switch_gst= round(n.issuer_wdl_approved * f["npci_switch_gst"], 3)
    wdl_txn_amt   = round(n.issuer_wdl_total_amount, 2)

    bi_fee        = round(n.issuer_bi_approved  * f["bi_fee"], 2)
    bi_fee_gst    = round(n.issuer_bi_approved  * f["bi_fee_gst"], 2)
    bi_switch     = round(n.issuer_bi_approved  * f["npci_switch"], 2)
    bi_switch_gst = round(n.issuer_bi_approved  * f["npci_switch_gst"], 3)

    ms_fee        = round(n.issuer_ms_approved  * f["ms_fee"], 2)
    ms_fee_gst    = round(n.issuer_ms_approved  * f["ms_fee_gst"], 2)
    ms_switch     = round(n.issuer_ms_approved  * f["npci_switch"], 2)
    ms_switch_gst = round(n.issuer_ms_approved  * f["npci_switch_gst"], 3)

    pc_fee        = round(n.issuer_pc_approved  * f["pc_fee"], 2)
    pc_fee_gst    = round(n.issuer_pc_approved  * f["pc_fee_gst"], 2)
    pc_switch     = round(n.issuer_pc_approved  * f["npci_switch"], 2)
    pc_switch_gst = round(n.issuer_pc_approved  * f["npci_switch_gst"], 3)

    # Late reversals → CREDIT (money returned to bank)
    rev_fee_cr    = round(n.issuer_late_reversals * f["reversal_fee"], 2)
    rev_fee_gst_cr= round(n.issuer_late_reversals * f["reversal_fee_gst"], 2)
    rev_txn_cr    = round(n.issuer_late_rev_amount, 2)

    # Total debit (bank → NFS)
    total_debit = round(
        wdl_fee + wdl_fee_gst + wdl_switch + wdl_switch_gst + wdl_txn_amt +
        bi_fee  + bi_fee_gst  + bi_switch  + bi_switch_gst  +
        ms_fee  + ms_fee_gst  + ms_switch  + ms_switch_gst  +
        pc_fee  + pc_fee_gst  + pc_switch  + pc_switch_gst, 2)

    # Total credit (NFS → bank)
    total_credit = round(rev_fee_cr + rev_fee_gst_cr + rev_txn_cr, 2)

    # Dispute adjustments
    cb_debit  = round(n.chargeback_debit, 2)
    cb_credit = round(n.chargeback_credit, 2)
    net_adj   = round(cb_credit - cb_debit, 2)

    settlement_amount = round(total_debit - total_credit, 2)
    final_amount = round(settlement_amount - net_adj, 2) if net_adj > 0 else settlement_amount

    return {
        "line_items": [
            # Description, count, debit, credit
            # ── Issuer WDL ─────────────────────────
            ("Issuer WDL Approved Fee",
             n.issuer_wdl_approved, wdl_fee, 0),
            ("Issuer WDL Approved Fee - GST",
             n.issuer_wdl_approved, wdl_fee_gst, 0),
            ("Issuer WDL Approved NPCI Switching Fee",
             n.issuer_wdl_approved, wdl_switch, 0),
            ("Issuer WDL Approved NPCI Switching Fee - GST",
             n.issuer_wdl_approved, wdl_switch_gst, 0),
            ("Issuer WDL Transaction Amount",
             n.issuer_wdl_approved, wdl_txn_amt, 0),
            ("Issuer WDL Declined",
             n.issuer_wdl_declined, 0, 0),
            # Late reversals (credits back to bank)
            ("Issuer WDL - Processed Late Reversals and Reversed Issuer WDL Approved fee",
             n.issuer_late_reversals, 0, rev_fee_cr),
            ("Issuer WDL - Processed Late Reversals and Reversed Issuer WDL Approved fee - GST",
             n.issuer_late_reversals, 0, rev_fee_gst_cr),
            ("Issuer WDL - Processed Late Reversals and Reversed Issuer WDL Transaction Amount",
             n.issuer_late_reversals, 0, rev_txn_cr),
            # ── Issuer BI ──────────────────────────
            ("Issuer BI Approved Fee",
             n.issuer_bi_approved, bi_fee, 0),
            ("Issuer BI Approved Fee - GST",
             n.issuer_bi_approved, bi_fee_gst, 0),
            ("Issuer BI Approved NPCI Switching Fee",
             n.issuer_bi_approved, bi_switch, 0),
            ("Issuer BI Approved NPCI Switching Fee - GST",
             n.issuer_bi_approved, bi_switch_gst, 0),
            ("Issuer BI Declined",
             n.issuer_bi_declined, 0, 0),
            # ── Issuer MS ──────────────────────────
            ("Issuer MS Approved Fee",
             n.issuer_ms_approved, ms_fee, 0),
            ("Issuer MS Approved Fee - GST",
             n.issuer_ms_approved, ms_fee_gst, 0),
            ("Issuer MS Approved NPCI Switching Fee",
             n.issuer_ms_approved, ms_switch, 0),
            ("Issuer MS Approved NPCI Switching Fee - GST",
             n.issuer_ms_approved, ms_switch_gst, 0),
            ("Issuer MS Declined",
             n.issuer_ms_declined, 0, 0),
            # ── Issuer PC ──────────────────────────
            ("Issuer PC Approved Fee",
             n.issuer_pc_approved, pc_fee, 0),
            ("Issuer PC Approved Fee - GST",
             n.issuer_pc_approved, pc_fee_gst, 0),
            ("Issuer PC Approved NPCI Switching Fee",
             n.issuer_pc_approved, pc_switch, 0),
            ("Issuer PC Approved NPCI Switching Fee - GST",
             n.issuer_pc_approved, pc_switch_gst, 0),
            ("Issuer PC Declined",
             n.issuer_pc_declined, 0, 0),
        ],
        "total_debit":        total_debit,
        "total_credit":       total_credit,
        "settlement_amount":  settlement_amount,
        "cb_debit":           cb_debit,
        "cb_credit":          cb_credit,
        "net_adj":            net_adj,
        "final_amount":       final_amount,
    }


# ─────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────

# Colours matching the real NFS settlement file
C_HEADER_BG   = "1F4E79"   # dark blue header
C_HEADER_FG   = "FFFFFF"   # white text
C_SECTION_BG  = "D6E4F0"   # light blue section rows
C_SUBTOTAL_BG = "BDD7EE"   # medium blue subtotals
C_TOTAL_BG    = "2E75B6"   # blue for final totals
C_TOTAL_FG    = "FFFFFF"
C_ALT_ROW     = "EBF3FB"   # alternating row light blue
C_ZERO        = "808080"   # grey for zero amounts
C_DEBIT       = "C00000"   # red tint for debit
C_CREDIT      = "375623"   # green tint for credit
C_DISPUTE     = "F4B942"   # amber dispute section


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border() -> Border:
    thin = Side(style="thin", color="B8CCE4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=False)

def _num(val: float) -> str:
    """Format number for display: blank if zero, 2 dp otherwise."""
    if val == 0:
        return ""
    return f"{val:,.2f}"


def write_settlement_xlsx(stats: SettlementStats, out_path: Path) -> Path:
    """Write the NFS Daily Settlement Statement Excel file."""

    computed = compute_line_items(stats)
    items    = computed["line_items"]
    date_str = stats.settlement_date.strftime("%d/%m/%Y") \
               if stats.settlement_date else datetime.today().strftime("%d/%m/%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement"

    # Column widths
    ws.column_dimensions["A"].width = 68
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    row = 1

    # ── Bank header ──────────────────────────────────────────
    ws.cell(row, 1, f"Daily Settlement Statement for {stats.bank_name} as on {date_str}")
    ws.cell(row, 1).font   = _font(bold=True, color=C_HEADER_FG, size=11)
    ws.cell(row, 1).fill   = _fill(C_HEADER_BG)
    ws.cell(row, 1).alignment = _align()
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    # Blank row
    row += 1

    # ── Column headers ────────────────────────────────────────
    for col, hdr in enumerate(["Description", "No of Txns", "Debit", "Credit"], 1):
        c = ws.cell(row, col, hdr)
        c.font      = _font(bold=True, color=C_HEADER_FG)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _align("center")
        c.border    = _border()
    row += 1

    # ── Line items ────────────────────────────────────────────
    alt = False
    for i, (desc, count, debit, credit) in enumerate(items):
        # Skip zero-count rows with no amounts (keep declined zero rows though)
        if count == 0 and debit == 0 and credit == 0:
            continue

        bg = C_ALT_ROW if alt else "FFFFFF"
        alt = not alt

        # Description
        c = ws.cell(row, 1, desc)
        c.font      = _font()
        c.fill      = _fill(bg)
        c.alignment = _align()
        c.border    = _border()

        # Count
        c = ws.cell(row, 2, count if count else "")
        c.font      = _font()
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()

        # Debit
        dv = debit if debit else ""
        c = ws.cell(row, 3, dv)
        c.font      = _font(color=C_DEBIT if debit else C_ZERO)
        c.fill      = _fill(bg)
        c.alignment = _align("right")
        c.border    = _border()
        if isinstance(dv, float) and dv > 0:
            c.number_format = "#,##0.00"

        # Credit
        cv = credit if credit else ""
        c = ws.cell(row, 4, cv)
        c.font      = _font(color=C_CREDIT if credit else C_ZERO)
        c.fill      = _fill(bg)
        c.alignment = _align("right")
        c.border    = _border()
        if isinstance(cv, float) and cv > 0:
            c.number_format = "#,##0.00"

        row += 1

    # ── Settlement Charges ────────────────────────────────────
    row += 1
    ws.cell(row, 1, "Settlement Charges").font = _font()
    ws.cell(row, 3, 0).font = _font()
    row += 1

    # ── Subtotals ─────────────────────────────────────────────
    for col in range(1, 5):
        ws.cell(row, col).fill = _fill(C_SUBTOTAL_BG)
        ws.cell(row, col).border = _border()
    ws.cell(row, 1, "Issuer/ Acquirer Sub Totals").font = _font(bold=True)
    ws.cell(row, 1).fill = _fill(C_SUBTOTAL_BG)
    ws.cell(row, 3, computed["total_debit"]).number_format = "#,##0.00"
    ws.cell(row, 3).font = _font(bold=True, color=C_DEBIT)
    ws.cell(row, 3).fill = _fill(C_SUBTOTAL_BG)
    ws.cell(row, 4, computed["total_credit"] or "").number_format = "#,##0.00"
    ws.cell(row, 4).font = _font(bold=True, color=C_CREDIT)
    ws.cell(row, 4).fill = _fill(C_SUBTOTAL_BG)
    row += 1

    # ── Settlement Amount ─────────────────────────────────────
    row += 1
    ws.cell(row, 1, "Settlement Amount").font = _font(bold=True)
    ws.cell(row, 3, computed["settlement_amount"]).number_format = "#,##0.00"
    ws.cell(row, 3).font = _font(bold=True, color=C_DEBIT)
    ws.cell(row, 4, 0)
    row += 1

    # Net Adjusted (if chargebacks)
    row += 2
    ws.cell(row, 1, "Net Adjusted Amount").font = _font()
    ws.cell(row, 3, computed["cb_debit"] or 0)
    ws.cell(row, 4, computed["net_adj"] or 0)
    row += 1

    # ── Final Settlement ──────────────────────────────────────
    row += 1
    for col in range(1, 5):
        ws.cell(row, col).fill = _fill(C_TOTAL_BG)
        ws.cell(row, col).border = _border()
    ws.cell(row, 1, "Final Settlement Amount").font = _font(bold=True, color=C_TOTAL_FG)
    ws.cell(row, 1).fill = _fill(C_TOTAL_BG)
    ws.cell(row, 3, computed["final_amount"]).number_format = "#,##0.00"
    ws.cell(row, 3).font = _font(bold=True, color=C_TOTAL_FG)
    ws.cell(row, 3).fill = _fill(C_TOTAL_BG)
    row += 1

    # ── Dispute Adjustments ───────────────────────────────────
    row += 1
    ws.cell(row, 1, "Dispute Adjustments").font = _font(bold=True, color=C_HEADER_FG)
    ws.cell(row, 1).fill = _fill(C_DISPUTE)
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    row += 1
    for col, hdr in enumerate(["Description", "No of Txns", "Debit", "Credit"], 1):
        c = ws.cell(row, col, hdr)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_DISPUTE)
        c.border = _border()
    row += 1

    if stats.chargeback_count > 0:
        ws.cell(row, 1, "Total Chargeback Amount").font = _font()
        ws.cell(row, 2, stats.chargeback_count)
        ws.cell(row, 3, stats.chargeback_debit)
        ws.cell(row, 4, stats.chargeback_credit)
        row += 1

    # Adjustment subtotals
    row += 1
    for col in range(1, 5):
        ws.cell(row, col).fill = _fill(C_SUBTOTAL_BG)
    ws.cell(row, 1, "Adjustment Sub Totals").font = _font(bold=True)
    ws.cell(row, 1).fill = _fill(C_SUBTOTAL_BG)
    ws.cell(row, 3, computed["cb_debit"])
    ws.cell(row, 4, computed["cb_credit"])
    row += 1

    row += 1
    ws.cell(row, 1, "Net Adjusted Amount").font = _font()
    ws.cell(row, 4, computed["net_adj"] or "")
    row += 1

    out_path = Path(out_path)
    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────
# PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────────

def generate_settlement(
    manifest: dict,
    bank_name: str,
    output_dir: Path,
) -> str:
    """
    Generate a NFS settlement Excel file from a matrix manifest.

    Parameters
    ----------
    manifest   : the manifest dict produced by generate_matrix()
    bank_name  : name of the bank (shown in header)
    output_dir : directory to write the file

    Returns
    -------
    str path to the written .xlsx file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    stats = stats_from_manifest(manifest, bank_name)

    date_str = stats.settlement_date.strftime("%d%m%Y") \
               if stats.settlement_date else datetime.today().strftime("%d%m%Y")
    run_id   = manifest.get("run_id", datetime.now().strftime("%Y%m%d%H%M%S"))

    out_path = output_dir / f"NFS_SETTLEMENT_{date_str}_{run_id}.xlsx"
    write_settlement_xlsx(stats, out_path)

    computed = compute_line_items(stats)
    return {
        "path":               str(out_path),
        "bank_name":          bank_name,
        "settlement_date":    date_str,
        "issuer_wdl_approved": stats.issuer_wdl_approved,
        "issuer_wdl_amount":   stats.issuer_wdl_total_amount,
        "issuer_bi_approved":  stats.issuer_bi_approved,
        "total_debit":         computed["total_debit"],
        "total_credit":        computed["total_credit"],
        "final_settlement":    computed["final_amount"],
    }


if __name__ == "__main__":
    import json, sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from generators.matrix_generator import generate_matrix

    result = generate_matrix(volume=500, ok_pct=99.0)
    import json
    manifest_path = result["manifest_path"]
    manifest = json.loads(Path(manifest_path).read_text())

    out = generate_settlement(
        manifest=manifest,
        bank_name="TEST BANK LTD",
        output_dir=Path(__file__).parent.parent / "output",
    )
    print(f"Settlement file: {out['path']}")
    print(f"  WDL approved:     {out['issuer_wdl_approved']} txns")
    print(f"  WDL amount:       ₹{out['issuer_wdl_amount']:,.2f}")
    print(f"  Total Debit:      ₹{out['total_debit']:,.2f}")
    print(f"  Total Credit:     ₹{out['total_credit']:,.2f}")
    print(f"  Final Settlement: ₹{out['final_settlement']:,.2f}")
